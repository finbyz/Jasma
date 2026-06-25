import frappe
from frappe.utils import flt,getdate
import openpyxl
from openpyxl import Workbook


def update_balance(employee_advance):
    if not employee_advance:
        return

    doc = frappe.get_doc("Employee Advance", employee_advance)

    balance_amount = (
        flt(doc.paid_amount)
        - flt(doc.claimed_amount)
        - flt(doc.return_amount)
    )

    doc.db_set(
        "balance_amount",
        balance_amount,
        update_modified=False
    )


def update_employee_advance_balance(doc, method=None):

    # Expense Claim
    if doc.doctype == "Expense Claim":
        for row in doc.get("advances", []):
            update_balance(row.employee_advance)

    # Payment Entry
    elif doc.doctype == "Payment Entry":
        for row in doc.get("references", []):
            if row.reference_doctype == "Employee Advance":
                update_balance(row.reference_name)

    # Journal Entry
    elif doc.doctype == "Journal Entry":
        for row in doc.get("accounts", []):
            if row.reference_type == "Employee Advance":
                update_balance(row.reference_name)
                
                

import frappe
import json
import os
from frappe.utils import getdate
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


@frappe.whitelist()
def export_payment_entries(payment_entries):
    if isinstance(payment_entries, str):
        payment_entries = json.loads(payment_entries)

    wb = Workbook()
    ws = wb.active
    ws.title = "Payment Entries"

    headers = [
        "PYMT_PROD_TYPE_CODE",
        "PYMT_MODE",
        "DEBIT_ACC_NO",
        "BNF_NAME",
        "BENE_ACC_NO",
        "BENE_IFSC",
        "AMOUNT",
        "DEBIT_NARR",
        "CREDIT_NARR",
        "MOBILE_NUM",
        "EMAIL_ID",
        "REMARK",
        "PYMT_DATE",
        "REF_NO",
        "ADDL_INFO1",
        "ADDL_INFO2",
        "ADDL_INFO3",
        "ADDL_INFO4",
        "ADDL_INFO5"
    ]

    ws.append(headers)

    # Header Style
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for pe_name in payment_entries:
        pe = frappe.get_doc("Payment Entry", pe_name)

        party_name = pe.party_name or ""
        mode_of_payment = pe.mode_of_payment or ""
        amount = pe.paid_amount

        party_bank_account = ""
        bank_account = ""
        ifsc = ""
        
        if pe.bank_account:
            bank = frappe.get_doc("Bank Account", pe.bank_account)
            bank_account = bank.bank_account_no or ""
            ifsc = bank.ifs_code or ""


        if pe.party_bank_account:
            bank = frappe.get_doc("Bank Account", pe.party_bank_account)
            party_bank_account = bank.bank_account_no or ""

        row = [
            "PAB_VENDOR",
            mode_of_payment,
            bank_account,
            party_name,
            party_bank_account,
            ifsc,
            amount,
            pe.narration or "",
            "",
            "",
            "",
            pe.remarks or "",
            getdate(pe.posting_date),
            pe.reference_no or "",
            "",
            "",
            "",
            "",
            ""
        ]

        ws.append(row)

    # Apply border and alignment
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Column widths
    widths = {
        "A": 20,
        "B": 15,
        "C": 22,
        "D": 35,
        "E": 22,
        "F": 18,
        "G": 15,
        "H": 35,
        "I": 35,
        "J": 15,
        "K": 30,
        "L": 25,
        "M": 15,
        "N": 45,
        "O": 20,
        "P": 20,
        "Q": 20,
        "R": 20,
        "S": 20,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Format Amount Column
    for cell in ws["G"][1:]:
        cell.number_format = '#,##0.00'

    # Format Date Column
    for cell in ws["M"][1:]:
        cell.number_format = 'DD-MM-YYYY'

    # Freeze Header Row
    ws.freeze_panes = "A2"

    # Add Filter
    ws.auto_filter.ref = ws.dimensions

    # Create Table
    table = Table(
        displayName="PaymentEntryTable",
        ref=f"A1:S{ws.max_row}"
    )

    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )

    table.tableStyleInfo = style
    ws.add_table(table)

    # Save File
    file_name = "Payment_Entry_Export.xlsx"
    file_path = f"/tmp/{file_name}"

    wb.save(file_path)

    with open(file_path, "rb") as f:
        file_doc = frappe.get_doc({
            "doctype": "File",
            "file_name": file_name,
            "is_private": 0,
            "content": f.read()
        })
        file_doc.save(ignore_permissions=True)

    return file_doc.file_url